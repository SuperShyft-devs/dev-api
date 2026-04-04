"""Integration tests for employee routes (employee-only)."""

from __future__ import annotations

from datetime import timedelta
from io import BytesIO

import pytest

from core.config import settings
from core.security import create_jwt_token
from openpyxl import load_workbook

from modules.employee.models import Employee
from modules.users.models import User


def _auth_header(user_id: int) -> dict[str, str]:
    token = create_jwt_token({"sub": str(user_id)}, timedelta(minutes=5), secret_key=settings.JWT_SECRET_KEY)
    return {"Authorization": f"Bearer {token}"}


async def _seed_admin_employee(test_db_session, *, user_id: int, employee_id: int = 1):
    test_db_session.add(User(user_id=user_id, phone=f"{user_id}000000000", age=30, status="active"))
    await test_db_session.flush()  # Ensure user is inserted before employee
    test_db_session.add(Employee(employee_id=employee_id, user_id=user_id, role="admin", status="active"))
    await test_db_session.commit()


@pytest.mark.asyncio
async def test_database_backup_returns_xlsx_for_employee(async_client, test_db_session):
    await _seed_admin_employee(test_db_session, user_id=8010, employee_id=501)

    response = await async_client.get("/employees/database-backup", headers=_auth_header(8010))
    assert response.status_code == 200
    assert "spreadsheetml" in (response.headers.get("content-type") or "")
    assert "attachment" in (response.headers.get("content-disposition") or "").lower()

    wb = load_workbook(BytesIO(response.content))
    names = {n.lower() for n in wb.sheetnames}
    assert "users" in names
    assert "employee" in names


@pytest.mark.asyncio
async def test_database_backup_accepts_access_token_query(async_client, test_db_session):
    await _seed_admin_employee(test_db_session, user_id=8011, employee_id=502)

    token = create_jwt_token({"sub": "8011"}, timedelta(minutes=5), secret_key=settings.JWT_SECRET_KEY)
    response = await async_client.get(f"/employees/database-backup?access_token={token}")
    assert response.status_code == 200
    wb = load_workbook(BytesIO(response.content))
    assert len(wb.sheetnames) >= 1


@pytest.mark.asyncio
async def test_database_backup_rejects_non_employee(async_client, test_db_session):
    test_db_session.add(User(user_id=8012, phone="8012000000", age=30, status="active"))
    await test_db_session.commit()

    response = await async_client.get("/employees/database-backup", headers=_auth_header(8012))
    assert response.status_code == 403


@pytest.mark.asyncio
async def test_create_employee_requires_auth(async_client):
    response = await async_client.post("/employees", json={"user_id": 1, "role": "admin"})
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_create_employee_requires_employee(async_client, test_db_session):
    test_db_session.add(User(user_id=8001, phone="8001000000", age=30, status="active"))
    await test_db_session.commit()

    response = await async_client.post("/employees", headers=_auth_header(8001), json={"user_id": 2, "role": "admin"})
    assert response.status_code == 403


@pytest.mark.asyncio
async def test_create_employee_creates_row(async_client, test_db_session):
    await _seed_admin_employee(test_db_session, user_id=8002, employee_id=11)
    test_db_session.add(User(user_id=9000, phone="9000000000", age=30, status="active"))
    await test_db_session.commit()

    payload = {"user_id": 9000, "role": "admin"}
    response = await async_client.post("/employees", headers=_auth_header(8002), json=payload)
    assert response.status_code == 201

    employee_id = response.json()["data"]["employee_id"]
    created = await test_db_session.get(Employee, employee_id)
    assert created is not None
    assert created.user_id == 9000
    assert (created.status or "").lower() == "active"


@pytest.mark.asyncio
async def test_list_employees_paginates_and_filters(async_client, test_db_session):
    await _seed_admin_employee(test_db_session, user_id=8003, employee_id=12)

    test_db_session.add(User(user_id=9100, phone="9100000000", age=30, status="active"))
    test_db_session.add(User(user_id=9101, phone="9101000000", age=30, status="active"))
    await test_db_session.flush()
    test_db_session.add_all(
        [
            Employee(employee_id=100, user_id=9100, role="admin", status="active"),
            Employee(employee_id=101, user_id=9101, role="ops", status="inactive"),
        ]
    )
    await test_db_session.commit()

    response = await async_client.get("/employees?page=1&limit=10&status=active", headers=_auth_header(8003))
    assert response.status_code == 200

    body = response.json()
    assert body["meta"]["page"] == 1
    assert body["meta"]["limit"] == 10
    assert body["meta"]["total"] >= 1

    for row in body["data"]:
        assert (row["status"] or "").lower() == "active"


@pytest.mark.asyncio
async def test_get_employee_returns_details(async_client, test_db_session):
    await _seed_admin_employee(test_db_session, user_id=8004, employee_id=13)

    test_db_session.add(User(user_id=9201, phone="9201000000", age=30, status="active"))
    await test_db_session.flush()
    test_db_session.add(Employee(employee_id=201, user_id=9201, role="ops", status="active"))
    await test_db_session.commit()

    response = await async_client.get("/employees/201", headers=_auth_header(8004))
    assert response.status_code == 200
    assert response.json()["data"]["employee_id"] == 201
    assert response.json()["data"]["user_id"] == 9201


@pytest.mark.asyncio
async def test_update_employee_updates_fields(async_client, test_db_session):
    await _seed_admin_employee(test_db_session, user_id=8005, employee_id=14)

    test_db_session.add(User(user_id=9301, phone="9301000000", age=30, status="active"))
    await test_db_session.flush()
    test_db_session.add(Employee(employee_id=301, user_id=9301, role="ops", status="active"))
    await test_db_session.commit()

    payload = {"user_id": 9301, "role": "admin"}
    response = await async_client.put("/employees/301", headers=_auth_header(8005), json=payload)
    assert response.status_code == 200

    updated = await test_db_session.get(Employee, 301)
    assert updated is not None
    assert updated.role == "admin"


@pytest.mark.asyncio
async def test_update_employee_status_sets_inactive(async_client, test_db_session):
    await _seed_admin_employee(test_db_session, user_id=8006, employee_id=15)

    test_db_session.add(User(user_id=9401, phone="9401000000", age=30, status="active"))
    await test_db_session.flush()
    test_db_session.add(Employee(employee_id=401, user_id=9401, role="ops", status="active"))
    await test_db_session.commit()

    response = await async_client.patch(
        "/employees/401/status",
        headers=_auth_header(8006),
        json={"status": "inactive"},
    )
    assert response.status_code == 200

    updated = await test_db_session.get(Employee, 401)
    assert updated is not None
    assert (updated.status or "").lower() == "inactive"


@pytest.mark.asyncio
async def test_update_employee_status_rejects_inactive_for_employee_one(async_client, test_db_session):
    await _seed_admin_employee(test_db_session, user_id=8007, employee_id=15)

    test_db_session.add(User(user_id=9402, phone="9402000000", age=30, status="active"))
    await test_db_session.flush()
    test_db_session.add(Employee(employee_id=9500, user_id=9402, role="ops", status="active"))
    await test_db_session.commit()

    response = await async_client.patch(
        "/employees/9500/status",
        headers=_auth_header(8007),
        json={"status": "inactive"},
    )
    assert response.status_code == 400

    protected = await test_db_session.get(Employee, 9500)
    assert protected is not None
    assert (protected.status or "").lower() == "active"
